"""
utils/excel_exporter.py
Exports an import cost calculation to a formatted Excel file.
Requires: openpyxl
"""
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter

from models.cost_config import CostBreakdown, CostConfig, ExchangeRate
from models.product import ImportOrder
from utils.logger import get_logger

log = get_logger(__name__)

# Color palette
_C_HEADER  = "1E3A5F"   # dark blue
_C_SUBHEAD = "2D6A9F"   # medium blue
_C_COST    = "D32F2F"   # red — cost
_C_PROFIT  = "2E7D32"   # green — profit
_C_SELL    = "1565C0"   # blue — selling price
_C_ALT     = "F0F4FA"   # light blue alt row
_C_WHITE   = "FFFFFF"


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cur(num: float) -> str:
    return f"{num:,.0f}"


def export(
    order: ImportOrder,
    config: CostConfig,
    rate: ExchangeRate,
    breakdown: CostBreakdown,
    use_bank_rate: bool,
    out_path: Path,
) -> Path:
    """Generate a .xlsx report and save to out_path. Returns the path."""
    wb = openpyxl.Workbook()
    _write_products_sheet(wb.active, order, rate, use_bank_rate)
    _write_cost_sheet(wb.create_sheet("Chi phí & Lợi nhuận"), config, rate, breakdown, use_bank_rate)
    wb.save(out_path)
    log.info("Excel exported to %s", out_path)
    return out_path


def _write_products_sheet(ws, order: ImportOrder, rate: ExchangeRate, use_bank_rate: bool):
    ws.title = "Danh sách sản phẩm"
    ex = rate.bank_rate if use_bank_rate else rate.market_rate

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"BÁO GIÁ NHẬP KHẨU — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = _font(bold=True, color=_C_WHITE, size=14)
    ws["A1"].fill = _fill(_C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Rate info
    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Loại tiền: {order.currency}  |  Tỷ giá: "
                f"{'Ngân hàng' if use_bank_rate else 'Thị trường'}  |  "
                f"1 {order.currency} = {_cur(ex)} VND")
    ws["A2"].font = _font(color=_C_SUBHEAD)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["STT", "Tên sản phẩm", "Số lượng", f"Đơn giá ({order.currency})",
               f"Thành tiền ({order.currency})", "Tỷ giá", "Thành tiền (VND)", ""]
    widths  = [6, 35, 12, 18, 20, 14, 22, 5]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = _font(bold=True, color=_C_WHITE)
        cell.fill = _fill(_C_SUBHEAD)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    # Data rows
    for i, line in enumerate(order.lines, 1):
        r = 3 + i
        fill = _fill(_C_ALT) if i % 2 == 0 else _fill(_C_WHITE)
        vals = [
            i,
            line.product.name,
            line.product.qty,
            line.product.unit_price_foreign,
            line.total_foreign,
            ex,
            line.total_vnd,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill = fill
            cell.border = _border()
            if col in (3, 4, 5, 7):
                cell.number_format = "#,##0.00" if col in (3, 4, 5) else "#,##0"
            if col == 6:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right" if col > 2 else "left",
                                       vertical="center")

    # Totals row
    r_total = 3 + len(order.lines) + 1
    ws.cell(row=r_total, column=1, value="TỔNG")
    ws.merge_cells(f"A{r_total}:B{r_total}")
    t_foreign = ws.cell(row=r_total, column=5, value=order.total_foreign)
    t_vnd     = ws.cell(row=r_total, column=7, value=order.total_vnd)
    for cell in [ws.cell(row=r_total, column=c) for c in range(1, 8)]:
        cell.font = _font(bold=True)
        cell.fill = _fill(_C_HEADER[:-1] + "44" if len(_C_HEADER) == 6 else "E8F0FF")
        cell.border = _border()
    t_foreign.number_format = "#,##0.00"
    t_vnd.number_format     = "#,##0"
    ws.freeze_panes = "A4"


def _write_cost_sheet(ws, config: CostConfig, rate: ExchangeRate,
                      bd: CostBreakdown, use_bank_rate: bool):
    ws.title = "Chi phí & Lợi nhuận"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"] = "PHÂN TÍCH CHI PHÍ NHẬP KHẨU"
    ws["A1"].font = _font(bold=True, color=_C_WHITE, size=13)
    ws["A1"].fill = _fill(_C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    rows = [
        ("", "", ""),
        ("CHI PHÍ ĐẦU VÀO", "VND", "%"),
        ("Trị giá hàng hóa (FOB)", bd.total_vnd_base, "—"),
        ("  + Thuế nhập khẩu",     bd.import_tax_vnd, f"{config.import_tax_pct}%"),
        ("  + VAT",                bd.vat_vnd,         f"{config.vat_pct}%"),
        ("  + Phí chuyển đổi ngoại tệ", bd.fx_fee_vnd, f"{config.fx_conversion_pct}%"),
        ("  + Lệ phí hải quan",    bd.customs_fee_vnd,     "Fixed"),
        ("  + VAT lệ phí HQ",      bd.customs_fee_vat_vnd, f"{config.customs_fee_vat_pct}%"),
        ("", "", ""),
        ("GIÁ VỐN (TOTAL COST)",   bd.total_cost_vnd,    ""),
        ("", "", ""),
        ("ĐỊNH GIÁ BÁN", "", ""),
        ("  Margin",                config.margin_pct, "%"),
        ("  Giá bán đề xuất",       bd.selling_price_vnd, ""),
        ("  Lợi nhuận",             bd.profit_vnd, ""),
    ]

    for r_idx, (label, value, note) in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=2, value=value if isinstance(value, (int, float)) else "")
        ws.cell(row=r_idx, column=3, value=note)

        for c in range(1, 4):
            cell = ws.cell(row=r_idx, column=c)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="right" if c == 2 else "left",
                                       vertical="center")
            if isinstance(value, (int, float)) and c == 2:
                cell.number_format = "#,##0"

        # Highlight key rows
        if label == "GIÁ VỐN (TOTAL COST)":
            for c in range(1, 4):
                ws.cell(row=r_idx, column=c).fill = _fill("FFCDD2")
                ws.cell(row=r_idx, column=c).font = _font(bold=True, color=_C_COST)
        elif label == "  Giá bán đề xuất":
            for c in range(1, 4):
                ws.cell(row=r_idx, column=c).fill = _fill("BBDEFB")
                ws.cell(row=r_idx, column=c).font = _font(bold=True, color=_C_SELL)
        elif label == "  Lợi nhuận":
            for c in range(1, 4):
                ws.cell(row=r_idx, column=c).fill = _fill("C8E6C9")
                ws.cell(row=r_idx, column=c).font = _font(bold=True, color=_C_PROFIT)
        elif label in ("CHI PHÍ ĐẦU VÀO", "ĐỊNH GIÁ BÁN"):
            for c in range(1, 4):
                ws.cell(row=r_idx, column=c).fill = _fill(_C_SUBHEAD)
                ws.cell(row=r_idx, column=c).font = _font(bold=True, color=_C_WHITE)

    ws.freeze_panes = "A3"

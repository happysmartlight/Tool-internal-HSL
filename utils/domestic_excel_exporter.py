"""
utils/domestic_excel_exporter.py
Exports a domestic pricing calculation to a formatted Excel file (2 sheets).
  Sheet 1: "Nội bộ doanh nghiệp" — full cost/margin detail
  Sheet 2: "Báo giá khách hàng"  — clean selling price only (no cost info)
Requires: openpyxl
"""
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.domestic_product import DomesticBreakdown, DomesticCostConfig
from utils.logger import get_logger

log = get_logger(__name__)

# ── Color palette (mirrors excel_exporter.py) ────────────────
_C_HEADER  = "1E3A5F"   # dark blue
_C_SUBHEAD = "2D6A9F"   # medium blue
_C_COST    = "D32F2F"   # red — cost
_C_PROFIT  = "2E7D32"   # green — profit
_C_SELL    = "1565C0"   # blue — selling price
_C_MARGIN  = "6A1B9A"   # purple — margin
_C_ALT     = "F0F4FA"   # light blue alt row
_C_WHITE   = "FFFFFF"
_C_GRAY    = "F5F5F5"

SELLER_NAME = "CÔNG TY TNHH THƯƠNG MẠI VÀ CÔNG NGHỆ HAPPY SMART LIGHT"
SELLER_TAX  = "MST: 3502535621"


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cur(num: float) -> str:
    return f"{num:,.0f}"


def _pct(num: float) -> str:
    return f"{num:.1f}%"


def _apply(cell, bold=False, color="000000", size=10, fill=None,
           halign="center", valign="center", wrap=False):
    cell.font = _font(bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = _border()
    if fill:
        cell.fill = _fill(fill)


def export(
    breakdown: DomesticBreakdown,
    config: DomesticCostConfig,
    out_path: Path,
) -> Path:
    """Generate a .xlsx report and save to out_path. Returns the path."""
    wb = openpyxl.Workbook()
    _write_internal_sheet(wb.active, breakdown, config)
    _write_customer_sheet(wb.create_sheet("Báo giá khách hàng"), breakdown)
    wb.save(out_path)
    log.info("Domestic Excel exported to %s", out_path)
    return out_path


# ── Sheet 1: Nội bộ doanh nghiệp ─────────────────────────────

def _write_internal_sheet(ws, breakdown: DomesticBreakdown, config: DomesticCostConfig):
    ws.title = "Nội bộ doanh nghiệp"
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── Title ─────────────────────────────────────────────────
    NCOLS = 14
    last_col = get_column_letter(NCOLS)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"BÁO GIÁ NỘI ĐỊA - NỘI BỘ — {now_str}"
    _apply(ws["A1"], bold=True, color=_C_WHITE, size=14, fill=_C_HEADER)
    ws.row_dimensions[1].height = 30

    # ── Sub-header ────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    vat_info = f"VAT bán: {config.vat_on_sale_pct:.0f}%  |  Ship tổng: {_cur(config.shipping_total_vnd)} ₫  |  CP cố định: {_cur(config.other_fixed_costs_vnd)} ₫"
    ws["A2"] = vat_info
    _apply(ws["A2"], color=_C_WHITE, size=10, fill=_C_SUBHEAD)
    ws.row_dimensions[2].height = 20

    # ── Headers row 3 ─────────────────────────────────────────
    headers = [
        "STT", "Tên sản phẩm", "ĐVT", "SL",
        "Giá mua/đv (₫)", "Ship/đv (₫)", "CP khác/đv (₫)",
        "Giá vốn/đv (₫)", "Biên LN%",
        "Giá bán\n(chưa VAT) (₫)", "Giá bán\n(có VAT) (₫)",
        "LN/đv (₫)", "Tổng giá bán (₫)", "Tổng LN (₫)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        _apply(cell, bold=True, color=_C_WHITE, size=10, fill=_C_HEADER, wrap=True)
    ws.row_dimensions[3].height = 32

    # ── Data rows ─────────────────────────────────────────────
    for i, line in enumerate(breakdown.lines):
        r = 4 + i
        row_fill = _C_ALT if i % 2 == 0 else _C_WHITE
        vals = [
            i + 1,
            line.product_name,
            line.unit,
            line.qty,
            line.purchase_price_vnd,
            line.allocated_ship_vnd,
            line.allocated_other_vnd,
            line.unit_cost_vnd,
            line.margin_pct_used / 100,
            line.unit_sell_before_vat_vnd,
            line.unit_sell_with_vat_vnd,
            line.profit_per_unit_vnd,
            line.total_sell_with_vat_vnd,
            line.total_profit_vnd,
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            is_name = col_idx == 2
            _apply(cell, fill=row_fill,
                   halign="left" if is_name else "right",
                   size=10)
            # Number format
            if col_idx in (5, 6, 7, 8, 10, 11, 12, 13, 14):
                cell.number_format = '#,##0'
            elif col_idx == 9:
                cell.number_format = '0.0%'
            elif col_idx == 4:
                cell.number_format = '#,##0.##'

    # ── Summary block ─────────────────────────────────────────
    n = len(breakdown.lines)
    sr = 4 + n + 1  # summary start row

    summary_rows = [
        ("Tổng giá vốn (Giá vốn)",  breakdown.total_cost_vnd,            _C_COST,    _C_WHITE),
        ("Doanh thu (chưa VAT)",     breakdown.total_revenue_before_vat,  _C_SELL,    _C_WHITE),
        ("Thuế VAT",                 breakdown.vat_amount_vnd,            _C_SUBHEAD, _C_WHITE),
        ("Doanh thu (có VAT)",       breakdown.total_revenue_with_vat,    _C_SELL,    _C_WHITE),
        ("Lợi nhuận",                breakdown.total_profit_vnd,          _C_PROFIT,  _C_WHITE),
        (f"Biên LN trung bình",      None,                                _C_MARGIN,  _C_WHITE),
    ]

    LABEL_END = get_column_letter(NCOLS - 2)
    for j, (label, value, bg, fg) in enumerate(summary_rows):
        r = sr + j
        ws.merge_cells(f"A{r}:{LABEL_END}{r}")
        lbl_cell = ws[f"A{r}"]
        lbl_cell.value = label
        _apply(lbl_cell, bold=True, color=fg, size=11, fill=bg, halign="right")

        val_cell_col = get_column_letter(NCOLS - 1)
        val2_cell_col = get_column_letter(NCOLS)

        if label.startswith("Biên"):
            # Span last 2 cols with percentage
            ws.merge_cells(f"{val_cell_col}{r}:{val2_cell_col}{r}")
            vc = ws[f"{val_cell_col}{r}"]
            vc.value = breakdown.avg_margin_pct / 100
            vc.number_format = "0.0%"
            _apply(vc, bold=True, color=fg, size=11, fill=bg)
        else:
            ws.merge_cells(f"{val_cell_col}{r}:{val2_cell_col}{r}")
            vc = ws[f"{val_cell_col}{r}"]
            vc.value = value
            vc.number_format = "#,##0"
            _apply(vc, bold=True, color=fg, size=11, fill=bg)

    # ── Column widths ─────────────────────────────────────────
    col_widths = [5, 28, 6, 6, 16, 12, 12, 16, 9, 16, 16, 14, 18, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"


# ── Sheet 2: Báo giá khách hàng ──────────────────────────────

def _write_customer_sheet(ws, breakdown: DomesticBreakdown):
    ws.title = "Báo giá khách hàng"
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── Company header ────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = SELLER_NAME
    _apply(ws["A1"], bold=True, color=_C_WHITE, size=12, fill=_C_HEADER)
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:F2")
    ws["A2"] = SELLER_TAX
    _apply(ws["A2"], color=_C_WHITE, size=10, fill=_C_SUBHEAD)
    ws.row_dimensions[2].height = 18

    # ── Title ─────────────────────────────────────────────────
    ws.merge_cells("A3:F3")
    ws["A3"] = f"BẢNG BÁO GIÁ SẢN PHẨM — {now_str}"
    _apply(ws["A3"], bold=True, color=_C_WHITE, size=13, fill=_C_HEADER)
    ws.row_dimensions[3].height = 28

    # ── Headers ───────────────────────────────────────────────
    headers = ["STT", "Tên sản phẩm", "ĐVT", "Số lượng",
               "Đơn giá (VND, có VAT)", "Thành tiền (VND)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        _apply(cell, bold=True, color=_C_WHITE, size=11, fill=_C_HEADER)
    ws.row_dimensions[4].height = 22

    # ── Data rows ─────────────────────────────────────────────
    for i, line in enumerate(breakdown.lines):
        r = 5 + i
        row_fill = _C_ALT if i % 2 == 0 else _C_WHITE
        vals = [
            i + 1,
            line.product_name,
            line.unit,
            line.qty,
            line.unit_sell_with_vat_vnd,
            line.total_sell_with_vat_vnd,
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            is_name = col_idx == 2
            _apply(cell, fill=row_fill,
                   halign="left" if is_name else "right",
                   size=11)
            if col_idx in (5, 6):
                cell.number_format = '#,##0'
            elif col_idx == 4:
                cell.number_format = '#,##0.##'

    # ── Total row ─────────────────────────────────────────────
    n = len(breakdown.lines)
    tr = 5 + n
    ws.merge_cells(f"A{tr}:E{tr}")
    total_label = ws[f"A{tr}"]
    total_label.value = "TỔNG CỘNG"
    _apply(total_label, bold=True, color=_C_WHITE, size=12, fill=_C_COST, halign="right")

    total_val = ws.cell(row=tr, column=6, value=breakdown.total_revenue_with_vat)
    total_val.number_format = "#,##0"
    _apply(total_val, bold=True, color=_C_WHITE, size=12, fill=_C_COST)

    ws.row_dimensions[tr].height = 22

    # ── VAT note ──────────────────────────────────────────────
    note_r = tr + 1
    ws.merge_cells(f"A{note_r}:F{note_r}")
    note_cell = ws[f"A{note_r}"]
    note_cell.value = "* Giá trên đã bao gồm VAT. Báo giá có hiệu lực trong vòng 15 ngày kể từ ngày lập."
    note_cell.font = _font(bold=False, color="666666", size=9)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths ─────────────────────────────────────────
    col_widths = [5, 34, 7, 8, 22, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
